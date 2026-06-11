import csv
import io
import os
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response, current_app
from app import db
from sqlalchemy import and_, or_
from app.models import Stage, Entry, CompetitionItem, EventConfig, Participant, GroupEntry, StagePlanItem


def _competition_entry_filter():
    """SQLAlchemy filter that excludes individual tracking entries for group items."""
    return or_(
        and_(CompetitionItem.item_type == 'group',  Entry.group_id.isnot(None)),
        and_(CompetitionItem.item_type != 'group',  Entry.participant_id.isnot(None)),
    )


def _item_entries(item_id, stage_id=None, **extra_filters):
    """Return only scoreable/schedulable entries for an item."""
    item = CompetitionItem.query.get(item_id)
    q = Entry.query.filter_by(item_id=item_id, **extra_filters)
    if item and item.item_type == 'group':
        q = q.filter(Entry.group_id.isnot(None))
    else:
        q = q.filter(Entry.participant_id.isnot(None))
    if stage_id is not None:
        q = q.filter_by(stage_id=stage_id)
    return q

schedule_bp = Blueprint('schedule', __name__)

@schedule_bp.before_request
def require_admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('auth.login', next=request.path))


@schedule_bp.before_request
def require_admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('auth.login', next=request.path))


@schedule_bp.route('/')
def index():
    stages = Stage.query.order_by(Stage.display_order).all()
    unassigned = (Entry.query
                  .join(Entry.competition_item)
                  .filter(Entry.stage_id.is_(None))
                  .filter(_competition_entry_filter())
                  .order_by(Entry.id).all())
    return render_template('schedule/index.html', stages=stages, unassigned=unassigned)


@schedule_bp.route('/assign', methods=['POST'])
def assign():
    entry_id = int(request.form['entry_id'])
    stage_id = request.form.get('stage_id')
    entry = Entry.query.get_or_404(entry_id)
    if stage_id:
        entry.stage_id = int(stage_id)
        # Set running order at end of stage
        last = db.session.query(db.func.max(Entry.running_order)).filter_by(
            stage_id=int(stage_id)
        ).scalar() or 0
        entry.running_order = last + 1
    else:
        entry.stage_id = None
        entry.running_order = None
    db.session.commit()
    flash('Stage assignment updated.', 'success')
    return redirect(url_for('schedule.index'))


@schedule_bp.route('/stage/<int:stage_id>')
def stage_view(stage_id):
    stage = Stage.query.get_or_404(stage_id)
    entries = (Entry.query
               .join(Entry.competition_item)
               .filter(Entry.stage_id == stage_id)
               .filter(_competition_entry_filter())
               .order_by(Entry.running_order).all())
    return render_template('schedule/stage.html', stage=stage, entries=entries)


@schedule_bp.route('/entry/<int:entry_id>/status', methods=['POST'])
def update_status(entry_id):
    entry = Entry.query.get_or_404(entry_id)
    new_status = request.form['status']
    if new_status in ('waiting', 'performing', 'completed'):
        entry.status = new_status
        db.session.commit()
    return redirect(request.referrer or url_for('schedule.index'))


def _contact_phone(entry):
    if entry.participant:
        return entry.participant.phone or ''
    if entry.group_entry and entry.group_entry.members:
        return entry.group_entry.members[0].phone or ''
    return ''


def _build_schedule_data(category=''):
    """Return schedule grouped by stage → event (planned order) → entries."""
    stages = Stage.query.order_by(Stage.display_order).all()
    schedule_data = []
    for stage in stages:
        plan_items = (StagePlanItem.query
                      .filter_by(stage_id=stage.id)
                      .order_by(StagePlanItem.display_order)
                      .all())
        if plan_items:
            ordered_item_ids = [pi.item_id for pi in plan_items]
        else:
            items = (CompetitionItem.query
                     .order_by(CompetitionItem.category, CompetitionItem.name)
                     .all())
            ordered_item_ids = [i.id for i in items]

        event_groups = []
        for item_id in ordered_item_ids:
            entries = (_item_entries(item_id, stage_id=stage.id)
                       .order_by(Entry.running_order)
                       .all())
            if category:
                entries = [e for e in entries if e.competition_item.category == category]
            if entries:
                event_groups.append({
                    'item': CompetitionItem.query.get(item_id),
                    'entries': entries,
                })

        if event_groups:
            schedule_data.append({'stage': stage, 'event_groups': event_groups})
    return schedule_data


@schedule_bp.route('/print')
def print_schedule():
    category = request.args.get('category', '')
    categories = ['Kids', 'Sub-Junior', 'Junior', 'Senior', 'Super Senior', 'Common']
    cfg = EventConfig.query.first()
    schedule_data = _build_schedule_data(category)

    return render_template('schedule/print.html',
                           schedule_data=schedule_data,
                           category=category,
                           categories=categories,
                           cfg=cfg)


@schedule_bp.route('/export/csv')
def export_csv():
    category = request.args.get('category', '')
    schedule_data = _build_schedule_data(category)
    cfg = EventConfig.query.first()
    event_name = cfg.event_name if cfg else 'Kalamela'

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Stage', 'Event', 'Category', 'Type', 'Order', 'Chest #', 'Name / Group', 'Contact'])
    for row in schedule_data:
        for grp in row['event_groups']:
            item = grp['item']
            for e in grp['entries']:
                writer.writerow([
                    row['stage'].name,
                    item.name,
                    item.category,
                    item.item_type,
                    e.running_order,
                    e.chest_number,
                    e.display_name,
                    _contact_phone(e),
                ])

    output = buf.getvalue()
    response = make_response(output)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    filename = f"{event_name.replace(' ', '_')}_schedule.csv"
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@schedule_bp.route('/export/excel')
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is required for Excel export. Run: pip install openpyxl', 'danger')
        return redirect(url_for('schedule.print_schedule'))

    category = request.args.get('category', '')
    schedule_data = _build_schedule_data(category)
    cfg = EventConfig.query.first()
    event_name = cfg.event_name if cfg else 'Kalamela'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'

    # Title row
    ws.merge_cells('A1:H1')
    ws['A1'] = event_name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header row
    headers = ['Stage', 'Event', 'Category', 'Type', 'Order', 'Chest #', 'Name / Group', 'Contact']
    header_fill = PatternFill('solid', fgColor='1A1A2E')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    event_fill = PatternFill('solid', fgColor='DBE9FF')
    stage_fill = PatternFill('solid', fgColor='E9ECEF')

    current_row = 3
    for row in schedule_data:
        # Stage divider row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        cell = ws.cell(row=current_row, column=1, value=row['stage'].name)
        cell.fill = stage_fill
        cell.font = Font(bold=True)
        current_row += 1

        for grp in row['event_groups']:
            item = grp['item']
            # Event sub-header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            cell = ws.cell(row=current_row, column=1,
                           value=f"{item.name}  |  {item.category}  |  {item.item_type}")
            cell.fill = event_fill
            cell.font = Font(italic=True)
            current_row += 1

            for e in grp['entries']:
                ws.cell(row=current_row, column=1, value=row['stage'].name)
                ws.cell(row=current_row, column=2, value=item.name)
                ws.cell(row=current_row, column=3, value=item.category)
                ws.cell(row=current_row, column=4, value=item.item_type)
                ws.cell(row=current_row, column=5, value=e.running_order)
                ws.cell(row=current_row, column=6, value=e.chest_number)
                ws.cell(row=current_row, column=7, value=e.display_name)
                ws.cell(row=current_row, column=8, value=_contact_phone(e))
                current_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.getvalue())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{event_name.replace(' ', '_')}_schedule.xlsx"
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@schedule_bp.route('/chest-numbers/header', methods=['POST'])
def chest_numbers_header():
    from werkzeug.utils import secure_filename
    cfg = EventConfig.query.first()
    if not cfg:
        cfg = EventConfig()
        db.session.add(cfg)
        db.session.commit()

    f = request.files.get('chest_header_image')
    if f and f.filename:
        ext = f.filename.rsplit('.', 1)[-1].lower()
        if ext in {'png', 'jpg', 'jpeg'}:
            folder = current_app.config['UPLOAD_FOLDER']
            os.makedirs(folder, exist_ok=True)
            filename = secure_filename(f'chest_number_header.{ext}')
            f.save(os.path.join(folder, filename))
            cfg.chest_number_header_image = filename
            db.session.commit()
            flash('Header image saved.', 'success')
        else:
            flash('Only PNG or JPG files are accepted.', 'warning')

    if request.form.get('remove_header'):
        cfg.chest_number_header_image = None
        db.session.commit()
        flash('Header image removed.', 'success')

    return redirect(url_for('schedule.chest_numbers'))


@schedule_bp.route('/chest-numbers')
def chest_numbers():
    cfg = EventConfig.query.first()

    include_registered = request.args.get('registered', '1') == '1'
    range_from  = request.args.get('from',      type=int, default=0)
    range_to    = request.args.get('to',         type=int, default=0)
    font_size   = request.args.get('font_size',  type=int, default=0) or None
    form_submitted  = 'from' in request.args
    show_name       = ('show_name'   in request.args) if form_submitted else True
    show_header     = ('show_header' in request.args) if form_submitted else True
    header_h_mm     = request.args.get('header_h', type=int, default=25) or 25

    # Resolve header image
    upload_folder   = current_app.config.get('UPLOAD_FOLDER', '')
    header_img_path = None
    header_img_url  = None
    if cfg and cfg.chest_number_header_image:
        p = os.path.join(upload_folder, cfg.chest_number_header_image)
        if os.path.exists(p):
            header_img_url  = url_for('main.uploads', filename=cfg.chest_number_header_image)
            if show_header:
                header_img_path = p

    # Build name lookup for registered numbers
    participants = {p.chest_number: p.full_name
                    for p in Participant.query.all()}
    groups = {g.chest_number: g.group_name
              for g in GroupEntry.query.all()}
    registered_numbers = sorted(set(participants) | set(groups))

    numbers = []
    if include_registered:
        for n in registered_numbers:
            numbers.append({
                'number': n,
                'name': participants.get(n) or groups.get(n, ''),
                'registered': True,
            })

    # Extra range (skip already included)
    existing = {item['number'] for item in numbers}
    if range_from and range_to and range_from <= range_to:
        for n in range(range_from, range_to + 1):
            if n not in existing:
                numbers.append({'number': n, 'name': '', 'registered': False})

    numbers.sort(key=lambda x: x['number'])

    return render_template(
        'schedule/chest_numbers.html',
        numbers=numbers,
        registered_count=len(registered_numbers),
        include_registered=include_registered,
        range_from=range_from or '',
        range_to=range_to or '',
        font_size=font_size or '',
        show_name=show_name,
        show_header=show_header,
        header_h_mm=header_h_mm,
        header_img_url=header_img_url,
        cfg=cfg,
    )


@schedule_bp.route('/chest-numbers/pdf')
def chest_numbers_pdf():
    include_registered = request.args.get('registered', '1') == '1'
    range_from  = request.args.get('from',      type=int, default=0)
    range_to    = request.args.get('to',         type=int, default=0)
    font_size   = request.args.get('font_size',  type=int, default=0) or None
    show_name    = request.args.get('show_name',   '1') == '1'
    show_header  = request.args.get('show_header', '1') == '1'
    header_h_mm  = request.args.get('header_h',   type=int, default=25) or 25

    cfg = EventConfig.query.first()
    event_name = cfg.event_name if cfg else 'Kalamela'

    upload_folder   = current_app.config.get('UPLOAD_FOLDER', '')
    header_img_path = None
    if show_header and cfg and cfg.chest_number_header_image:
        p = os.path.join(upload_folder, cfg.chest_number_header_image)
        if os.path.exists(p):
            header_img_path = p

    participants = {p.chest_number: p.full_name for p in Participant.query.all()}
    groups = {g.chest_number: g.group_name for g in GroupEntry.query.all()}
    registered_numbers = sorted(set(participants) | set(groups))

    numbers = []
    if include_registered:
        for n in registered_numbers:
            numbers.append({
                'number': n,
                'name': participants.get(n) or groups.get(n, ''),
                'registered': True,
            })

    existing = {item['number'] for item in numbers}
    if range_from and range_to and range_from <= range_to:
        for n in range(range_from, range_to + 1):
            if n not in existing:
                numbers.append({'number': n, 'name': '', 'registered': False})

    numbers.sort(key=lambda x: x['number'])

    from app.pdf.chest_numbers import generate_chest_numbers_pdf
    pdf_bytes = generate_chest_numbers_pdf(numbers, event_name,
                                           font_size=font_size,
                                           header_img_path=header_img_path,
                                           header_h_mm=header_h_mm,
                                           show_name=show_name)

    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename="chest_numbers.pdf"'
    return response


@schedule_bp.route('/entry/<int:entry_id>/reorder', methods=['POST'])
def reorder(entry_id):
    entry = Entry.query.get_or_404(entry_id)
    direction = request.form.get('direction')
    if not entry.stage_id:
        return redirect(url_for('schedule.index'))

    siblings = Entry.query.filter_by(stage_id=entry.stage_id).order_by(Entry.running_order).all()
    idx = next((i for i, e in enumerate(siblings) if e.id == entry_id), None)
    if idx is None:
        return redirect(url_for('schedule.stage_view', stage_id=entry.stage_id))

    if direction == 'up' and idx > 0:
        siblings[idx].running_order, siblings[idx - 1].running_order = (
            siblings[idx - 1].running_order, siblings[idx].running_order
        )
    elif direction == 'down' and idx < len(siblings) - 1:
        siblings[idx].running_order, siblings[idx + 1].running_order = (
            siblings[idx + 1].running_order, siblings[idx].running_order
        )
    db.session.commit()
    return redirect(url_for('schedule.stage_view', stage_id=entry.stage_id))
