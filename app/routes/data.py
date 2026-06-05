import os
import shutil
import json
import csv
import io
from datetime import datetime
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, send_file, current_app, session)
from werkzeug.utils import secure_filename
from app import db
from app.models import (Participant, GroupEntry, Entry, Score,
                        CompetitionItem, Stage, EventConfig)

data_bp = Blueprint('data', __name__)

@data_bp.before_request
def require_admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('auth.login', next=request.path))


@data_bp.before_request
def require_admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('auth.login', next=request.path))


def _db_path():
    return current_app.config['DATABASE_PATH']


def _backup_folder():
    return current_app.config['BACKUP_FOLDER']


def _list_backups():
    folder = _backup_folder()
    files = [f for f in os.listdir(folder) if f.endswith('.db')]
    files.sort(reverse=True)
    return files


def _check_data_password(submitted):
    """Return True if the extra data-management password matches, or if none is configured."""
    required = current_app.config.get('DATA_RESET_PASSWORD', '')
    if not required:
        return True   # not configured — local dev, skip check
    return submitted == required


def _do_backup(label='manual'):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_{label}_{timestamp}.db'
    dest = os.path.join(_backup_folder(), filename)
    shutil.copy2(_db_path(), dest)
    return filename


@data_bp.route('/')
def index():
    backups = _list_backups()
    data_password_required = bool(current_app.config.get('DATA_RESET_PASSWORD'))
    return render_template('data/index.html', backups=backups,
                           data_password_required=data_password_required)


@data_bp.route('/backup', methods=['POST'])
def backup():
    filename = _do_backup(label='manual')
    flash(f'Backup created: {filename}', 'success')
    return redirect(url_for('data.index'))


@data_bp.route('/download/<filename>')
def download_backup(filename):
    safe = secure_filename(filename)
    path = os.path.join(_backup_folder(), safe)
    if not os.path.exists(path):
        flash('Backup file not found.', 'danger')
        return redirect(url_for('data.index'))
    return send_file(path, as_attachment=True, download_name=safe)


@data_bp.route('/restore', methods=['POST'])
def restore():
    if 'backup_file' not in request.files:
        flash('No file selected.', 'danger')
        return redirect(url_for('data.index'))

    file = request.files['backup_file']
    if not file.filename.endswith('.db'):
        flash('Only .db backup files are accepted.', 'danger')
        return redirect(url_for('data.index'))

    if not _check_data_password(request.form.get('data_password', '')):
        flash('Incorrect data management password.', 'danger')
        return redirect(url_for('data.index'))

    # Auto-backup before restore
    _do_backup(label='pre_restore')

    dest = _db_path()
    file.save(dest)
    flash('Database restored from backup. Refresh the page.', 'success')
    return redirect(url_for('data.index'))


@data_bp.route('/reset', methods=['POST'])
def reset():
    confirm = request.form.get('confirm_reset', '')
    if confirm != 'RESET':
        flash('Type RESET to confirm.', 'warning')
        return redirect(url_for('data.index'))

    if not _check_data_password(request.form.get('data_password', '')):
        flash('Incorrect data management password.', 'danger')
        return redirect(url_for('data.index'))

    # Auto-backup before reset
    backup_name = _do_backup(label='pre_reset')
    flash(f'Auto-backup created before reset: {backup_name}', 'info')

    # Delete participant/group/entry/score data — keep config, stages, items, criteria
    Score.query.delete()
    from app.models import AuditLog, group_members
    AuditLog.query.delete()
    db.session.execute(group_members.delete())
    Entry.query.delete()
    GroupEntry.query.delete()
    Participant.query.delete()
    db.session.commit()

    flash('All participant, group, entry, and score data has been cleared. '
          'Event settings, stages, and competition items are preserved.', 'success')
    return redirect(url_for('data.index'))


@data_bp.route('/delete-backup/<filename>', methods=['POST'])
def delete_backup(filename):
    safe = secure_filename(filename)
    path = os.path.join(_backup_folder(), safe)
    if os.path.exists(path):
        os.remove(path)
        flash(f'{safe} deleted.', 'success')
    return redirect(url_for('data.index'))


@data_bp.route('/export')
def export_data():
    format_ = request.args.get('format', 'json')

    participants = Participant.query.order_by(Participant.chest_number).all()
    groups = GroupEntry.query.order_by(GroupEntry.chest_number).all()
    entries = Entry.query.all()
    scores = Score.query.all()

    data = {
        'participants': [
            {
                'chest_number': p.chest_number,
                'full_name': p.full_name,
                'dob': p.date_of_birth.isoformat(),
                'category': p.category,
                'lkc_id': p.lkc_id,
                'gender': p.gender,
                'phone': p.phone,
                'email': p.email,
                'parent_name': p.parent_name,
            }
            for p in participants
        ],
        'groups': [
            {
                'chest_number': g.chest_number,
                'group_name': g.group_name,
                'item': g.item.name,
                'category': g.item.category,
                'members': [m.full_name for m in g.members],
            }
            for g in groups
        ],
        'entries': [
            {
                'id': e.id,
                'chest_number': e.chest_number,
                'name': e.display_name,
                'item': e.competition_item.name,
                'category': e.competition_item.category,
                'stage': e.stage.name if e.stage else None,
                'running_order': e.running_order,
                'status': e.status,
                'final_score': e.final_score,
            }
            for e in entries
        ],
        'scores': [
            {
                'entry_id': s.entry_id,
                'judge_number': s.judge_number,
                'criteria': s.criteria.name,
                'marks': s.marks,
            }
            for s in scores
        ],
    }

    if format_ == 'json':
        buf = io.BytesIO(json.dumps(data, indent=2).encode())
        return send_file(buf, mimetype='application/json',
                         as_attachment=True, download_name='kalamela_export.json')

    # CSV — participants sheet
    csv_buf = io.StringIO()
    writer = csv.DictWriter(csv_buf, fieldnames=[
        'chest_number', 'full_name', 'dob', 'category', 'lkc_id',
        'gender', 'phone', 'email', 'parent_name',
    ])
    writer.writeheader()
    writer.writerows(data['participants'])
    return send_file(
        io.BytesIO(csv_buf.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name='kalamela_participants.csv',
    )


def _entries_rows():
    """Return (headers, rows) for all entries, ordered by chest number."""
    entries = Entry.query.order_by(Entry.id).all()
    entries.sort(key=lambda e: (e.chest_number or 0, e.id))
    headers = ['Chest #', 'Name / Group', 'Type', 'Event', 'Category',
               'Stage', 'Running Order', 'Status', 'Contact']
    rows = []
    for e in entries:
        if e.participant:
            contact = e.participant.phone or ''
            etype = 'solo'
        elif e.group_entry:
            contact = e.group_entry.members[0].phone if e.group_entry.members else ''
            etype = 'group'
        else:
            contact = ''
            etype = ''
        rows.append([
            e.chest_number,
            e.display_name,
            etype,
            e.competition_item.name,
            e.competition_item.category,
            e.stage.name if e.stage else '',
            e.running_order or '',
            e.status or '',
            contact,
        ])
    return headers, rows


@data_bp.route('/export/entries/csv')
def export_entries_csv():
    cfg = EventConfig.query.first()
    event_name = cfg.event_name if cfg else 'Kalamela'
    headers, rows = _entries_rows()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)

    response_data = io.BytesIO(buf.getvalue().encode('utf-8'))
    return send_file(response_data, mimetype='text/csv', as_attachment=True,
                     download_name=f"{event_name.replace(' ', '_')}_entries.csv")


@data_bp.route('/export/entries/excel')
def export_entries_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is required for Excel export.', 'danger')
        return redirect(url_for('data.index'))

    cfg = EventConfig.query.first()
    event_name = cfg.event_name if cfg else 'Kalamela'
    headers, rows = _entries_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Entries'

    # Title
    ws.merge_cells(f'A1:{chr(64 + len(headers))}1')
    ws['A1'] = f'{event_name} — All Entries'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header row
    header_fill = PatternFill('solid', fgColor='1A1A2E')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows with alternating shading
    alt_fill = PatternFill('solid', fgColor='F2F2F2')
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    col_widths = [10, 30, 8, 28, 14, 18, 14, 12, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{event_name.replace(' ', '_')}_entries.xlsx",
    )
